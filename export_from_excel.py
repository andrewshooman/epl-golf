import json

import openpyxl


SOURCE_FILE = r"C:\Users\sushi\Downloads\Premier League 25_26 Table Golf (1).xlsx"
SHEET_NAME = "OFFICIAL Table Golf 2526"
OUT_DIR = r"C:\Users\sushi\OneDrive\Desktop\epl-prediction-standings-app\data"


def detect_bonus_rule(formula: str) -> str:
    f = formula.replace(" ", "").lower()
    if "<6" in f:
        return "top5"
    if "=6" in f:
        return "sixth"
    if "=7" in f:
        return "seventh"
    if ">17" in f:
        return "bottom3"
    return "none"


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=False)
    ws = wb[SHEET_NAME]

    standings = [
        {"team": str(ws.cell(r, 1).value).strip(), "position": int(r - 1)}
        for r in range(2, 22)
        if ws.cell(r, 1).value
    ]

    predictions = []
    for col in range(1, ws.max_column + 1):
        player_name = ws.cell(22, col).value
        has_pick_column = ws.cell(2, col).value is not None
        has_score_column = ws.cell(2, col + 1).value is not None
        if not isinstance(player_name, str) or not has_pick_column or not has_score_column:
            continue

        for row in range(2, 22):
            team = ws.cell(row, col).value
            formula = ws.cell(row, col + 1).value
            if not team or not isinstance(formula, str):
                continue
            predictions.append(
                {
                    "player": player_name.strip(),
                    "predictedPosition": int(row - 1),
                    "team": str(team).strip(),
                    "bonusRule": detect_bonus_rule(formula),
                }
            )

    with open(f"{OUT_DIR}\\predictions.json", "w", encoding="utf-8") as f:
        json.dump(predictions, f, indent=2)
    with open(f"{OUT_DIR}\\fallback-standings.json", "w", encoding="utf-8") as f:
        json.dump(standings, f, indent=2)

    print("predictions:", len(predictions))
    print("players:", sorted({item["player"] for item in predictions}))


if __name__ == "__main__":
    main()
